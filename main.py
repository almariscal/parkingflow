import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime, timedelta
from tkcalendar import Calendar as TkCalendar  # Renombramos Calendar de tkcalendar
import os
from pathlib import Path
from openpyxl import load_workbook
import zipfile
from icalendar import Calendar as IcalCalendar, Event  # Renombramos Calendar de icalendar

# Función para leer los datos del Excel y obtener las citas dentro del rango de fechas
def leer_datos(excel_path, fecha_inicio_str, fecha_fin_str):
    wb = load_workbook(excel_path)
    ws_asignaciones = wb["Asignaciones"]
    ws_empleados = wb["Empleados"]

    # Construir el mapa de iniciales a (nombre, email)
    mapa_empleados = {}
    for row in ws_empleados.iter_rows(min_row=2, values_only=True):
        iniciales, nombre, email = row
        mapa_empleados[iniciales] = {'nombre': nombre, 'email': email}

    # Construir las citas
    plazas = list(ws_asignaciones.iter_rows(min_row=1, max_row=1, values_only=True))[0][1:]
    citas_por_empleado = {}

    fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
    fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()

    for row in ws_asignaciones.iter_rows(min_row=2, values_only=True):
        fecha_str = row[0]
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        if not (fecha_inicio <= fecha <= fecha_fin):
            continue
        asignaciones = row[1:]
        for idx, iniciales in enumerate(asignaciones):
            if iniciales not in mapa_empleados:
                continue
            plaza = plazas[idx]
            empleado = mapa_empleados[iniciales]
            nombre = empleado['nombre']
            email = empleado['email']
            evento = {
                'fecha': fecha,
                'plaza': plaza,
                'nombre': nombre,
                'email': email
            }
            citas_por_empleado.setdefault(iniciales, []).append(evento)

    return citas_por_empleado, mapa_empleados

# Función para generar los archivos .ics para cada empleado
def generar_archivos_ics(citas_por_empleado, mapa_empleados, output_dir):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    for iniciales, eventos in citas_por_empleado.items():
        carpeta_empleado = output_dir / iniciales
        carpeta_empleado.mkdir(exist_ok=True)
        for evento in eventos:
            cal = IcalCalendar()  # Crear un nuevo calendario para cada evento
            ev = Event()  # Crear el evento

            fecha = evento['fecha']
            ev.add('summary', f"PARKING - {evento['plaza']} - {evento['nombre']}")
            ev.add('dtstart', datetime.combine(fecha, datetime.strptime("08:00", "%H:%M").time()))
            ev.add('dtend', datetime.combine(fecha, datetime.strptime("08:05", "%H:%M").time()))  # Asignar duración
            ev.add('location', "Parking empresa")
            ev.add('status', "CONFIRMED")
            ev.add('description', "Reserva automática de plaza de parking")
            ev.add('transp', "TRANSPARENT")
            ev.add('uid', f"{evento['email']}@parking.com")

            cal.add_component(ev)  # Usar add_component en lugar de append o add

            # Escribir el archivo .ics
            file_name = f"{fecha.strftime('%Y-%m-%d')}.ics"
            with open(carpeta_empleado / file_name, 'w', encoding='utf-8') as f:
                f.write(cal.to_ical().decode('utf-8'))  # Serializamos correctamente a formato .ics

# Función para crear el archivo .zip con todos los .ics generados
def crear_zip(input_dir, zip_output):
    with zipfile.ZipFile(zip_output, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(input_dir):
            for file in files:
                zipf.write(os.path.join(root, file), os.path.relpath(os.path.join(root, file), input_dir))


# Función para abrir el explorador de archivos y seleccionar el archivo Excel
def seleccionar_excel():
    archivo = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
    if archivo:
        entry_excel.delete(0, tk.END)
        entry_excel.insert(0, archivo)

# Función para abrir el explorador de archivos y seleccionar la carpeta de salida
def seleccionar_carpeta():
    carpeta = filedialog.askdirectory()
    if carpeta:
        entry_carpeta.delete(0, tk.END)
        entry_carpeta.insert(0, carpeta)

# Función que se ejecuta al hacer click en el botón "Generar Citas"
def generar_citas():
    fecha_inicio = calendar_inicio.get_date()
    fecha_fin = calendar_fin.get_date()
    excel_path = entry_excel.get()
    carpeta_salida = entry_carpeta.get()

    if not excel_path or not carpeta_salida:
        messagebox.showerror("Error", "Por favor, seleccione un archivo Excel y una carpeta de salida.")
        return

    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_dir = Path(carpeta_salida) / f"salida_citas_{timestamp}"
        zip_output = Path(carpeta_salida) / f"citas_parking_{timestamp}.zip"
        
        # Llamamos a la función que genera los archivos de citas (debe estar definida previamente)
        citas_por_empleado, mapa_empleados = leer_datos(excel_path, fecha_inicio, fecha_fin)
        generar_archivos_ics(citas_por_empleado, mapa_empleados, temp_dir)
        crear_zip(temp_dir, zip_output)
        
        messagebox.showinfo("Completado", f"Las citas han sido generadas correctamente.\nZIP guardado en: {zip_output}")
    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error: {e}")

# Configuración de la interfaz gráfica
root = tk.Tk()
root.title("Generador de Citas de Parking")
root.geometry("450x500")

# Estilo y diseño
root.config(bg="#f0f0f0")

# Etiqueta para la fecha de inicio
label_inicio = tk.Label(root, text="Fecha de Inicio:", bg="#f0f0f0")
label_inicio.grid(row=0, column=0, padx=10, pady=10, sticky="w")

# Calendario para la fecha de inicio
calendar_inicio = TkCalendar(root, selectmode='day', date_pattern='yyyy-mm-dd')
calendar_inicio.selection_set((datetime.today() + timedelta(days=(7 - datetime.today().weekday()))).strftime('%Y-%m-%d'))  # Lunes de la semana que viene
calendar_inicio.grid(row=0, column=1, padx=10, pady=10)

# Etiqueta para la fecha de fin
label_fin = tk.Label(root, text="Fecha de Fin:", bg="#f0f0f0")
label_fin.grid(row=1, column=0, padx=10, pady=10, sticky="w")

# Calendario para la fecha de fin
calendar_fin = TkCalendar(root, selectmode='day', date_pattern='yyyy-mm-dd')
calendar_fin.selection_set((datetime.today() + timedelta(days=(7 + 4 - datetime.today().weekday()))).strftime('%Y-%m-%d'))  # Viernes de la semana que viene
calendar_fin.grid(row=1, column=1, padx=10, pady=10)

# Etiqueta para el archivo Excel
label_excel = tk.Label(root, text="Archivo Excel:", bg="#f0f0f0")
label_excel.grid(row=2, column=0, padx=10, pady=10, sticky="w")

# Entrada para el archivo Excel con valor predeterminado
entry_excel = tk.Entry(root, width=40)
entry_excel.grid(row=2, column=1, padx=10, pady=10)
entry_excel.insert(0, "asignacion_parking.xlsx")  # Valor predeterminado

# Botón para abrir el explorador de archivos para el Excel
boton_excel = tk.Button(root, text="Seleccionar Excel", command=seleccionar_excel)
boton_excel.grid(row=3, column=1, padx=10, pady=10, sticky="e")

# Etiqueta para la carpeta de salida
label_carpeta = tk.Label(root, text="Carpeta de Salida:", bg="#f0f0f0")
label_carpeta.grid(row=4, column=0, padx=10, pady=10, sticky="w")

# Entrada para la carpeta de salida con valor predeterminado (directorio actual)
entry_carpeta = tk.Entry(root, width=40)
entry_carpeta.grid(row=4, column=1, padx=10, pady=10)
entry_carpeta.insert(0, str(Path.cwd()))  # Carpeta actual por defecto

# Botón para abrir el explorador de archivos para la carpeta de salida
boton_carpeta = tk.Button(root, text="Seleccionar Carpeta", command=seleccionar_carpeta)
boton_carpeta.grid(row=5, column=1, padx=10, pady=10, sticky="e")

# Botón para generar las citas
boton_generar = tk.Button(root, text="Generar Citas", command=generar_citas, bg="#4CAF50", fg="white", font=("Arial", 12))
boton_generar.grid(row=6, column=0, columnspan=2, pady=20)

# Iniciar la interfaz gráfica
root.mainloop()
